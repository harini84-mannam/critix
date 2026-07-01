from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import admin as django_admin
from django.db.models import Count, Avg, Q
from django.apps import apps
from django.http import Http404, JsonResponse
from django.urls import reverse
from django.forms import modelform_factory
from django.views.decorators.http import require_POST

from .models import Movie, Review, ReviewReport, Watchlist, ReviewReply, ReportAction
from .models import Movie, Review, ReviewReport, Watchlist, ReviewReply, notify_user
from . import admin_metadata

def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_staff:
            return redirect('home')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


HIDDEN_APPS = set()

EXCLUDED_MODELS = set()


def _list_url_for(model):
    return reverse('admin_model_list', kwargs={'model_name': model._meta.model_name})


def _add_url_for(model):
    meta = admin_metadata.get_model_meta(model)
    editable = _editable_fields(model)
    if not editable:
        return None
    return reverse('admin_model_add', kwargs={'model_name': model._meta.model_name})


def get_apps_overview(request=None):

    apps_overview = []
    for app_config in apps.get_app_configs():
        if app_config.label in HIDDEN_APPS:
            continue
        app_models = []
        for model in sorted(app_config.get_models(), key=lambda m: m._meta.verbose_name.lower()):
            if (app_config.label, model._meta.model_name) in EXCLUDED_MODELS:
                continue
            app_models.append({
                'label':      model._meta.verbose_name_plural.title(),
                'model_name': model._meta.model_name,
                'count':      model.objects.count(),
                'list_url':   _list_url_for(model),
                'add_url':    _add_url_for(model),
            })
        apps_overview.append({
            'app_label': app_config.label,
            'app_name':  app_config.verbose_name,
            'models':    app_models,
            'is_open':   False,
        })

    if request is not None:
        for app in apps_overview:
            app['is_open'] = any(
                m['list_url'] and request.path.startswith(m['list_url'])
                for m in app['models']
            )

    return apps_overview

def _all_models():
    return sorted(apps.get_models(), key=lambda m: m._meta.verbose_name.lower())


def _get_model_by_name(model_name):
    for m in _all_models():
        if m._meta.model_name == model_name.lower():
            return m
    raise Http404(f'No model named "{model_name}".')


def _editable_fields(model):
    return [
        f.name for f in model._meta.fields
        if f.editable
        and not f.primary_key
        and not getattr(f, 'auto_now_add', False)
        and not getattr(f, 'auto_now', False)
    ]

@admin_required
def admin_dashboard(request):
    return render(request, 'admin_portal/dashboard.html', {})

@admin_required
def admin_metadata_api(request, model_name):
    model = _get_model_by_name(model_name)
    meta = admin_metadata.get_model_meta(model)
    safe_actions = [
        {'name': a['name'], 'label': a['label'], 'confirm': a['confirm']}
        for a in meta['actions']
    ]
    return JsonResponse({
        'model_name':    meta['model_name'],
        'verbose_name':  meta['verbose_name'],
        'list_display':  meta['list_display'],
        'search_fields': meta['search_fields'],
        'list_filter':   [{'name': f['name'], 'label': f['label']} for f in meta['list_filter']],
        'ordering':      meta['ordering'],
        'actions':       safe_actions,
    })

@admin_required
def admin_model_list(request, model_name):
    
    model = _get_model_by_name(model_name)
    meta  = admin_metadata.get_model_meta(model)
    qs = model.objects.all()
    query = request.GET.get('q', '').strip()
    if query and meta['search_fields']:
        q_obj = Q()
        for sf in meta['search_fields']:
            q_obj |= Q(**{f'{sf}__icontains': query})
        qs = qs.filter(q_obj)
    active_filters = {}
    filter_options = []
    for f in meta['list_filter']:
        selected = request.GET.get(f['name'], '')
        if selected:
            active_filters[f['name']] = selected
            try:
                qs = qs.filter(**{f['name']: selected})
            except Exception:
                pass
        filter_options.append({
            'name':     f['name'],
            'label':    f['label'],
            'choices':  f['choices'],
            'selected': selected,
        })
    sort_field = request.GET.get('sort', '')
    sort_dir   = request.GET.get('dir', 'asc')

    display_names = [col['name'] for col in meta['list_display']]

    if sort_field and sort_field in display_names:
        try:
            model._meta.get_field(sort_field)
            order_expr = f'-{sort_field}' if sort_dir == 'desc' else sort_field
            qs = qs.order_by(order_expr)
        except Exception:
            qs = qs.order_by(*meta['ordering'])
    else:
        qs = qs.order_by(*meta['ordering'])

    total_count   = qs.count()
    showing_capped = total_count > 300
    qs = qs[:300]

    rows = []
    for obj in qs:
        cells = [
            admin_metadata.resolve_display_value(obj, col['name'])
            for col in meta['list_display']
        ]
        rows.append({'pk': obj.pk, 'cells': cells})
    actions = [
        {'name': a['name'], 'label': a['label'], 'confirm': a['confirm']}
        for a in meta['actions']
    ]

    has_search = bool(meta['search_fields'])
    can_add    = bool(_editable_fields(model))

    context = {
        'model_name':     meta['model_name'],
        'verbose_name':   meta['verbose_name'],
        'verbose_name_pl': meta['verbose_name_pl'],
        'columns':        meta['list_display'],      
        'rows':           rows,                      
        'total_count':    total_count,
        'showing_capped': showing_capped,
        'filter_options': filter_options,
        'actions':        actions,
        'has_search':     has_search,
        'can_add':        can_add,
        'query':          query,
        'sort_field':     sort_field,
        'sort_dir':       sort_dir,
        'active_filters': active_filters,
    }
    return render(request, 'admin_portal/model_list.html', context)

@admin_required
@require_POST
def admin_model_action(request, model_name):
    
    model = _get_model_by_name(model_name)
    meta  = admin_metadata.get_model_meta(model)

    action_name  = request.POST.get('action', '')
    selected_raw = request.POST.get('selected_ids', '')

    action_entry = next(
        (a for a in meta['actions'] if a['name'] == action_name), None
    )
    if not action_entry:
        return redirect('admin_model_list', model_name=model_name)

    if selected_raw == 'all':
        qs = model.objects.all()
    else:
        pks = [pk.strip() for pk in selected_raw.split(',') if pk.strip()]
        qs  = model.objects.filter(pk__in=pks)

    if qs.exists():
        # If it's a delete action, redirect to confirmation page
        if action_name == 'bulk_delete' or (action_entry and 'delete' in action_name.lower()):
            ids_param = selected_raw
            return redirect(
                reverse('admin_bulk_delete_confirm', kwargs={'model_name': model_name})
                + f'?selected_ids={ids_param}'
            )
        model_admin = django_admin.site._registry.get(model)
        result = action_entry['_func'](model_admin, request, qs)
        if result is not None:
            return result

    return redirect('admin_model_list', model_name=model_name)

@admin_required
def admin_bulk_delete_confirm(request, model_name):
    """GET: show confirmation page. POST: perform the bulk delete."""
    model = _get_model_by_name(model_name)
    meta  = admin_metadata.get_model_meta(model)

    if request.method == 'POST':
        selected_raw = request.POST.get('selected_ids', '')
        if selected_raw == 'all':
            qs = model.objects.all()
        else:
            pks = [pk.strip() for pk in selected_raw.split(',') if pk.strip()]
            qs  = model.objects.filter(pk__in=pks)
        qs.delete()
        return redirect('admin_model_list', model_name=model_name)

    # GET — build confirmation info
    selected_raw = request.GET.get('selected_ids', '')
    if selected_raw == 'all':
        qs = model.objects.all()
    else:
        pks = [pk.strip() for pk in selected_raw.split(',') if pk.strip()]
        qs  = model.objects.filter(pk__in=pks)

    objects = list(qs[:50])  # cap display at 50

    from django.contrib.admin.utils import get_deleted_objects
    deleted_objects, model_count, perms_needed, protected = get_deleted_objects(
        list(qs), request, django_admin.site,
    )

    return render(request, 'admin_portal/confirm_delete.html', {
        'model_name':      meta['model_name'],
        'verbose_name':    meta['verbose_name'],
        'verbose_name_pl': meta['verbose_name_pl'],
        'objects':         objects,
        'total_count':     qs.count(),
        'selected_ids':    selected_raw,
        'deleted_objects': deleted_objects,
        'model_count':     model_count,
        'perms_needed':    perms_needed,
        'protected':       protected,
        'mode':            'bulk',
    })


@admin_required
def admin_model_add(request, model_name):
    model     = _get_model_by_name(model_name)
    FormClass = modelform_factory(model, fields=_editable_fields(model))

    if request.method == 'POST':
        form = FormClass(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save()
            if '_addanother' in request.POST:
                return redirect('admin_model_add', model_name=model_name)
            if '_continue' in request.POST:
                return redirect('admin_model_edit', model_name=model_name, pk=obj.pk)
            return redirect('admin_model_list', model_name=model_name)
    else:
        form = FormClass()

    meta = admin_metadata.get_model_meta(model)
    return render(request, 'admin_portal/model_form.html', {
        'form':         form,
        'verbose_name': meta['verbose_name'],
        'model_name':   model_name,
        'mode':         'add',
    })


@admin_required
def admin_model_edit(request, model_name, pk):
    model     = _get_model_by_name(model_name)
    obj       = get_object_or_404(model, pk=pk)
    FormClass = modelform_factory(model, fields=_editable_fields(model))

    if request.method == 'POST':
        form = FormClass(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            form.save()
            if '_addanother' in request.POST:
                return redirect('admin_model_add', model_name=model_name)
            if '_continue' in request.POST:
                return redirect('admin_model_edit', model_name=model_name, pk=pk)
            return redirect('admin_model_list', model_name=model_name)
    else:
        form = FormClass(instance=obj)

    meta = admin_metadata.get_model_meta(model)
    return render(request, 'admin_portal/model_form.html', {
        'form':         form,
        'verbose_name': meta['verbose_name'],
        'model_name':   model_name,
        'mode':         'edit',
        'pk':           pk,
    })


@admin_required
def admin_model_delete(request, model_name, pk):
    model = _get_model_by_name(model_name)
    obj   = get_object_or_404(model, pk=pk)
    meta  = admin_metadata.get_model_meta(model)

    if request.method == 'POST':
        obj.delete()
        return redirect('admin_model_list', model_name=model_name)

    # GET — show confirmation page
    # Collect related objects that will also be deleted
    from django.contrib.admin.utils import get_deleted_objects
    using = obj.__class__._default_manager.db
    deleted_objects, model_count, perms_needed, protected = get_deleted_objects(
        [obj], request, django_admin.site,
    )

    return render(request, 'admin_portal/confirm_delete.html', {
        'model_name':      meta['model_name'],
        'verbose_name':    meta['verbose_name'],
        'verbose_name_pl': meta['verbose_name_pl'],
        'object':          obj,
        'deleted_objects': deleted_objects,
        'model_count':     model_count,
        'perms_needed':    perms_needed,
        'protected':       protected,
        'mode':            'single',
    })

@admin_required
def admin_genres(request):
    all_movies = Movie.objects.values_list('genre', flat=True)
    genre_map = {}
    for raw in all_movies:
        for g in [x.strip() for x in raw.replace(',', '/').split('/') if x.strip()]:
            genre_map[g] = genre_map.get(g, 0) + 1
    genres  = sorted(genre_map.items(), key=lambda x: -x[1])
    success = request.session.pop('genre_success', None)
    error   = request.session.pop('genre_error', None)
    return render(request, 'admin_portal/genres.html', {
        'genres': genres, 'success': success, 'error': error,
    })


@admin_required
def admin_genre_rename(request):
    if request.method == 'POST':
        old_name = request.POST.get('old_name', '').strip()
        new_name = request.POST.get('new_name', '').strip()
        if not new_name:
            request.session['genre_error'] = 'New genre name cannot be empty.'
            return redirect('admin_genres')
        updated = 0
        for movie in Movie.objects.all():
            genres = [g.strip() for g in movie.genre.replace(',', '/').split('/')]
            if old_name in genres:
                genres = [new_name if g == old_name else g for g in genres]
                movie.genre = ' / '.join(genres)
                movie.save()
                updated += 1
        request.session['genre_success'] = (
            f'Renamed "{old_name}" → "{new_name}" on {updated} movie(s).'
        )
    return redirect('admin_genres')


@admin_required
def admin_genre_delete(request):
    if request.method == 'POST':
        genre_name = request.POST.get('genre_name', '').strip()
        updated = 0
        for movie in Movie.objects.all():
            genres = [g.strip() for g in movie.genre.replace(',', '/').split('/')]
            if genre_name in genres:
                genres = [g for g in genres if g != genre_name]
                movie.genre = ' / '.join(genres) if genres else 'Uncategorized'
                movie.save()
                updated += 1
        request.session['genre_success'] = (
            f'Removed genre "{genre_name}" from {updated} movie(s).'
        )
    return redirect('admin_genres')

import csv
import io
from django.http import HttpResponse

@admin_required
def admin_model_export(request, model_name):
    model   = _get_model_by_name(model_name)
    meta    = admin_metadata.get_model_meta(model)
    fmt     = request.GET.get('fmt', 'csv')  

    qs = model.objects.all()

    query = request.GET.get('q', '').strip()
    if query and meta['search_fields']:
        from django.db.models import Q
        q_obj = Q()
        for sf in meta['search_fields']:
            q_obj |= Q(**{f'{sf}__icontains': query})
        qs = qs.filter(q_obj)

    for f in meta['list_filter']:
        selected = request.GET.get(f['name'], '')
        if selected:
            try:
                qs = qs.filter(**{f['name']: selected})
            except Exception:
                pass

    qs = qs.order_by(*meta['ordering'])

    columns = meta['list_display']
    headers = [col['label'] for col in columns]

    rows = []
    for obj in qs:
        rows.append([
            admin_metadata.resolve_display_value(obj, col['name'])
            for col in columns
        ])

    if fmt == 'xlsx':
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = meta['verbose_name_pl']
            ws.append(headers)
            for row in rows:
                ws.append(row)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="{model_name}_export.xlsx"'
            return response
        except ImportError:
            pass

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{model_name}_export.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response

@admin_required
def admin_model_import(request, model_name):
    model = _get_model_by_name(model_name)
    meta  = admin_metadata.get_model_meta(model)
    session_key = f'import_{model_name}'
    request.session.pop(session_key, None)
    editable_names = _editable_fields(model)

    field_info = []
    for f in model._meta.fields:
        if f.name not in editable_names:
            continue
        info = {'name': f.name, 'type': f.get_internal_type(), 'fk_hint': None}
        if hasattr(f, 'related_model') and f.related_model:
            rm = f.related_model
            info['fk_hint'] = (
                f'FK → {rm._meta.verbose_name} '
                f'(use ID, or username/title/name)'
            )
            info['type'] = 'ForeignKey'
        field_info.append(info)

    return render(request, 'admin_portal/model_import.html', {
        'model_name':      meta['model_name'],
        'verbose_name':    meta['verbose_name'],
        'verbose_name_pl': meta['verbose_name_pl'],
        'field_info':      field_info,
    })


def _parse_import_file(uploaded):
    import csv, io as _io
    headers, rows, error = [], [], None
    filename = uploaded.name.lower()
    try:
        if filename.endswith('.csv'):
            text   = uploaded.read().decode('utf-8-sig')
            reader = csv.DictReader(_io.StringIO(text))
            headers = list(reader.fieldnames or [])
            for row in reader:
                rows.append({k: (v or '') for k, v in row.items()})
        elif filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb       = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
            ws       = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                headers = [str(h) if h is not None else '' for h in all_rows[0]]
                for row in all_rows[1:]:
                    rows.append({
                        headers[i]: (str(v) if v is not None else '')
                        for i, v in enumerate(row) if i < len(headers)
                    })
        else:
            error = 'Unsupported file type. Please upload .csv or .xlsx.'
    except ImportError:
        error = 'openpyxl is not installed. Run: pip install openpyxl'
    except Exception as e:
        error = f'Could not parse file: {e}'
    return headers, rows, error


def _normalize_headers(row, editable_field_names):
    field_map = {}
    for f in editable_field_names:
        field_map[f.lower()] = f
        field_map[f.replace('_', ' ').lower()] = f   
        field_map[f.replace('_', '').lower()] = f    
        if f.endswith('_id'):
            base = f[:-3]
            field_map[base.lower()] = f
            field_map[base.replace('_', ' ').lower()] = f

    normalized = {}
    for k, v in row.items():
        key = k.strip().lower().replace('-', ' ').replace('_', ' ')
        key_under = key.replace(' ', '_')
        if key_under in field_map:
            normalized[field_map[key_under]] = v
        elif key.replace(' ', '') in field_map:
            normalized[field_map[key.replace(' ', '')]] = v
        else:
            normalized[k] = v
    return normalized


def _resolve_fk_value(field, raw_value):
    
    if not raw_value or str(raw_value).strip() == '':
        return None

    val = str(raw_value).strip()
    related_model = field.related_model

    if val.isdigit():
        try:
            related_model.objects.get(pk=int(val))
            return int(val)
        except related_model.DoesNotExist:
            return None

    from django.contrib.auth.models import User
    if related_model == User:
        obj = User.objects.filter(username=val).first() or User.objects.filter(email=val).first()
        return obj.pk if obj else None

    for lookup_field in ('title', 'name', 'username', 'email', 'slug'):
        try:
            obj = related_model.objects.filter(**{lookup_field: val}).first()
            if obj:
                return obj.pk
        except Exception:
            pass

    for obj in related_model.objects.all()[:500]:
        if str(obj) == val:
            return obj.pk

    return None


def _do_import(model, rows):
    
    from django.forms import modelform_factory
    editable_names = _editable_fields(model)
    fk_fields = {}      
    regular_names = []  

    for f in model._meta.fields:
        if f.name not in editable_names and f.attname not in editable_names:
            continue
        if hasattr(f, 'related_model') and f.related_model is not None:
            fk_fields[f.attname] = f   
            fk_fields[f.name] = f      
        else:
            regular_names.append(f.name)

    FormClass = modelform_factory(model, fields=editable_names)

    for fname, form_field in FormClass.base_fields.items():
        from django import forms as _forms
        if isinstance(form_field, (_forms.CharField, _forms.IntegerField)):
            form_field.required = False

    created, errors = 0, []

    for i, row in enumerate(rows, 1):
        norm = _normalize_headers(row, editable_names + list(fk_fields.keys()))

        data = {}

        for fname in editable_names:
            if fname in fk_fields or fname + '_id' in fk_fields:
                continue  
            data[fname] = norm.get(fname, '')

        processed_fks = set()
        for f in model._meta.fields:
            if f.name not in editable_names:
                continue
            if not (hasattr(f, 'related_model') and f.related_model):
                continue
            if f.attname in processed_fks:
                continue
            processed_fks.add(f.attname)

            raw = norm.get(f.attname) or norm.get(f.name) or norm.get(f.name.replace('_', ' ')) or ''
            pk = _resolve_fk_value(f, raw)

            if pk is not None:
                data[f.attname] = pk   
            elif f.null:
                data[f.attname] = None  
            
        form = FormClass(data=data)
        if form.is_valid():
            form.save()
            created += 1
        else:
            flat = {k: ', '.join(v) for k, v in form.errors.items()}
            errors.append(f'Row {i}: {flat}')

    return created, errors


@admin_required
@require_POST
def admin_model_import_preview(request, model_name):
    model      = _get_model_by_name(model_name)
    meta       = admin_metadata.get_model_meta(model)
    action     = request.POST.get('action', 'preview')
    session_key = f'import_{model_name}'

    if action == 'direct':
        uploaded = request.FILES.get('import_file')
        if not uploaded:
            return redirect('admin_model_import', model_name=model_name)
        headers, rows, error = _parse_import_file(uploaded)
        if error:
            return render(request, 'admin_portal/model_import.html', {
                'model_name':      meta['model_name'],
                'verbose_name':    meta['verbose_name'],
                'verbose_name_pl': meta['verbose_name_pl'],
                'parse_error':     error,
            })
        created, errors = _do_import(model, rows)
        request.session.pop(session_key, None)
        return render(request, 'admin_portal/model_import_result.html', {
            'model_name':      meta['model_name'],
            'verbose_name':    meta['verbose_name'],
            'verbose_name_pl': meta['verbose_name_pl'],
            'created':         created,
            'errors':          errors,
        })

    if action == 'confirm':
        import json
        rows = request.session.get(session_key, [])
        if not rows:
            return redirect('admin_model_import', model_name=model_name)
        created, errors = _do_import(model, rows)
        request.session.pop(session_key, None)
        return render(request, 'admin_portal/model_import_result.html', {
            'model_name':      meta['model_name'],
            'verbose_name':    meta['verbose_name'],
            'verbose_name_pl': meta['verbose_name_pl'],
            'created':         created,
            'errors':          errors,
        })

    uploaded = request.FILES.get('import_file')
    if not uploaded:
        return redirect('admin_model_import', model_name=model_name)

    headers, rows, parse_error = _parse_import_file(uploaded)

    if not parse_error:
        request.session[session_key] = rows
        request.session.modified = True

    import json as _json
    preview = rows[:20]
    return render(request, 'admin_portal/model_import_preview.html', {
        'model_name':      meta['model_name'],
        'verbose_name':    meta['verbose_name'],
        'verbose_name_pl': meta['verbose_name_pl'],
        'headers':         headers,
        'preview_rows':    preview,
        'total_rows':      len(rows),
        'parse_error':     parse_error,
    })